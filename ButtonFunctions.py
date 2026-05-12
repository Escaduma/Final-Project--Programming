from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox

def add_placeholder(entry, text): # Adds placeholder text to an entry field that disappears when clicked
    entry.insert(0, text)
    entry.config(fg="grey")
    def on_focus_in(e): # Remove placeholder when user clicks the entry
        if entry.get() == text:
            entry.delete(0, END)
            entry.config(fg="black")
    def on_focus_out(e):
        if entry.get() == "": # Restore placeholder if user leaves the entry empty
            entry.insert(0, text)
            entry.config(fg="grey")
    entry.bind("<FocusIn>", on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

def clearRightPanel(rightPanel): # Clears all widgets from the right panel before loading new content
    for widget in rightPanel.winfo_children():
        widget.destroy()

def addBook(title, author,genre, year, description, read): # Saves a new book to the Excel file
    workbook=load_workbook("info/BooksList.xlsx")
    sheet=workbook.active
    status = "Read" if read == 1 else "Not Read"  # Convert checkbox value to Read/Not Read string
    row=[title, author, genre, year, description, status]
    sheet.append(row)
    workbook.save("info/BooksList.xlsx")
    messagebox.showinfo("Saved", "Your book has been saved!!")

def showAddBookForm(rightPanel): # Displays the Add Book form in the right panel
    clearRightPanel(rightPanel)
    bg = PhotoImage(file="info/add.png")# Background image
    bgLabel = Label(rightPanel, image=bg)
    bgLabel.place(x=0, y=0, relwidth=1, relheight=1)
    bgLabel.image = bg

    Label(rightPanel, text="Add a Book", font=("Brush Script MT", 25), bg="#FFE4E1").pack(ipady=10)

    # Title field
    titleLabel= Label(rightPanel, text="Book Title:", font=("Brush Script MT", 20), bg="#FFE4E1")
    titleLabel.pack(pady=(15,0))
    titleEntry = Entry(rightPanel, width=50,font=(None,14))
    titleEntry.pack()


    # Author field
    authorLabel= Label(rightPanel, text="Author:", font=("Brush Script MT", 20), bg="#FFE4E1")
    authorLabel.pack(pady=(10,0))
    authorEntry = Entry(rightPanel, width=50,font=(None,14))
    authorEntry.pack()

    #Genre Field
    genreLabel= Label(rightPanel, text="Genre:", font=("Brush Script MT", 20), bg="#FFE4E1")
    genreLabel.pack(pady=(10,0))
    genreEntry = Entry(rightPanel, width=50,font=(None,14))  
    genreEntry.pack()

    #Year Field
    yearLabel= Label(rightPanel, text="Year:", font=("Brush Script MT", 20), bg="#FFE4E1")
    yearLabel.pack(pady=(10,0))
    yearEntry = Entry(rightPanel, width=50,font=(None,14))
    yearEntry.pack()

    # Description field (multiple lines)
    descriptionLabel= Label(rightPanel, text="Description:", font=("Brush Script MT", 20), bg="#FFE4E1")
    descriptionLabel.pack(pady=(10,0))
    descriptionText = Text(rightPanel, width=50, height=5, font=(None,12))
    descriptionText.pack()

    # Checkbox to mark book as already read
    readVar = IntVar()
    Checkbutton(rightPanel, text="Already Read", variable=readVar, bg="#FFE4E1", font=("Brush Script MT", 18)).pack(pady=5)

    # Save button calls addBook with all field values
    def save():
        addBook(titleEntry.get(), authorEntry.get(), genreEntry.get(), yearEntry.get(), descriptionText.get("1.0", END), readVar.get())

    Button(rightPanel, text="  Save  ", bg="pink", command=save, font=("Brush Script MT", 17)).pack(pady=10)

def showRemoveBookForm(rightPanel): # Displays the Remove Book form in the right panel
    clearRightPanel(rightPanel)
    bg = PhotoImage(file="info/remove.png") # Background image
    bgLabel = Label(rightPanel, image=bg)
    bgLabel.place(x=0, y=0, relwidth=1, relheight=1)
    bgLabel.image = bg

    Label(rightPanel, text="Remove a Book", font=("Brush Script MT", 20), bg="#FFE4E1").pack(ipady=10)

    # Entry to type the title of the book to remove
    removeLabel= Label(rightPanel, text="Title of the book:", font=("Brush Script MT", 18), bg="#FFE4E1")
    removeLabel.pack()
    titleEntry = Entry(rightPanel, width=50,font=(None,16))
    titleEntry.pack(pady=(10,0))

    def remove():
        workbook = load_workbook("info/BooksList.xlsx")
        sheet = workbook.active
        target = titleEntry.get().strip().lower()
        removed = False
        # Search for the book by title and delete its row
        for row in sheet.iter_rows():
            if row[0].value and row[0].value.strip().lower() == target:
                sheet.delete_rows(row[0].row)
                removed = True
                break
        workbook.save("info/BooksList.xlsx")
        if removed:
            messagebox.showinfo("Removed", f'"{titleEntry.get()}" has been removed.')
        else:
            messagebox.showwarning("Not Found", f'"{titleEntry.get()}" was not found.')

    Button(rightPanel, text="  Remove  ", bg="pink", command=remove, font=("Brush Script MT", 18)).pack(pady=10)


def showRecommendation(rightPanel): # Displays the Recommendation form in the right panel
    clearRightPanel(rightPanel)
    bg = PhotoImage(file="info/recommendation.png") # Background image
    bgLabel = Label(rightPanel, image=bg)
    bgLabel.place(x=0, y=0, relwidth=1, relheight=1)
    bgLabel.image = bg

    Label(rightPanel, text="Book Recommendation", font=("Brush Script MT", 22), bg="#FFE4E1").pack(pady=10)

    # Genre filter entry with placeholder
    Label(rightPanel, text="Genre:", font=("Brush Script MT", 20), bg="#FFE4E1").pack()
    genreEntry = Entry(rightPanel, width=30, font=(None, 13))
    genreEntry.pack(pady=5, ipady=5)
    add_placeholder(genreEntry, "e.g. Fantasy, Romance...")

    # Radio buttons to filter by read status
    Label(rightPanel, text="Show me:", font=("Brush Script MT", 18), bg="#FFE4E1").pack(pady=(10,0))
    readVar = StringVar(value="Both")
    Frame_radio = Frame(rightPanel, bg="#FFE4E1")
    Frame_radio.pack()
    Radiobutton(Frame_radio, text="Not Read", variable=readVar, value="Not Read", bg="#FFE4E1", font=("Brush Script MT", 18)).pack(side="left", padx=5)
    Radiobutton(Frame_radio, text="Already Read", variable=readVar, value="Read", bg="#FFE4E1", font=("Brush Script MT", 18)).pack(side="left", padx=5)
    Radiobutton(Frame_radio, text="Both", variable=readVar, value="Both", bg="#FFE4E1", font=("Brush Script MT", 18)).pack(side="left", padx=5)

    # Frame where the recommendation result will be displayed
    resultFrame = Frame(rightPanel, bg="#FFE4E1")
    resultFrame.pack(pady=10)

    def recommend(): # Clear previous recommendation
        for widget in resultFrame.winfo_children():
            widget.destroy()

        workbook = load_workbook("info/BooksList.xlsx")
        sheet = workbook.active
        genre = genreEntry.get().strip().lower()
        status = readVar.get()

        # Filter books by genre and read status
        books = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            row_genre = str(row[2]).strip().lower() if row[2] else ""
            row_status = row[5] if row[5] else "Not Read"
            if genre and genre != "e.g. fantasy, romance..." and genre not in row_genre:
                continue
            if status != "Both" and row_status != status:
                continue
            books.append(row)

        if books: # Pick a random book from the filtered list
            import random
            pick = random.choice(books)
            Label(resultFrame, text=pick[0], font=("Brush Script MT", 20), bg="#FFE4E1").pack(pady=5)
            Label(resultFrame, text=pick[4], font=("Brush Script MT", 20), bg="#FFE4E1", wraplength=300, justify="center").pack(pady=5)
        else:
            Label(resultFrame, text="No books found!", font=("Brush Script MT", 14), bg="#FFE4E1").pack(pady=10)

    Button(rightPanel, text="  Recommend  ", bg="pink", font=("Brush Script MT", 20), command=recommend).pack(pady=10)
        
def showBookList(rightPanel): # Displays all saved books in a scrollable list
    clearRightPanel(rightPanel)

    Label(rightPanel, text="Your Books", font=("Brush Script MT", 22), bg="#FFE4E1").pack(ipady=10)

    workbook = load_workbook("info/BooksList.xlsx")
    sheet = workbook.active

    # Frame to hold the listbox and scrollbar
    frame = Frame(rightPanel, bg="#FFE4E1")
    frame.pack(fill="both", expand=True)

    scrollbar = Scrollbar(frame)
    scrollbar.pack(side="right", fill="y")

    # Listbox to display all books
    listbox = Listbox(frame, yscrollcommand=scrollbar.set, width=50, font=(None,16), height=15)
    listbox.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=listbox.yview)

    # Insert each book into the listbox
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if rows and any(r[0] for r in rows):
        for row in rows:
            if row[0]:
                status = row[5] if row[5] else "Not Read"
                listbox.insert(END, f"{row[0]} — {row[1]} ({row[2]}) [{row[3]}]  ---> {status}")  # Title — Author (Year) [Genre]
    else:
        listbox.insert(END, "No books saved yet.")